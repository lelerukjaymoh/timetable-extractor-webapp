#!/usr/bin/env python3
from openpyxl import Workbook
import openpyxl
import re
import json
from .models import Timetable, Year, Semester, Exam
from django.db.models import Q


class Extractor:
    
    def get_data(self, file_path, indentifier, timetable_id):
        # file = "Nov 2020 Exam timetable Y4 Y1(1).xlsx"
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        count = 1
        data = []
        date_start_row = 1
        date = ''

        for row in ws.iter_rows():
            for cell in row:
                if indentifier in str(cell.value):
                    # fourth.append(cell.value.split('–')[0].strip())
                    column_no = cell.column
                    row_no = cell.row
                    year = indentifier[1]
                    sem = indentifier[-1]
                    unit_name_unformatted = ws.cell(column=int(column_no)+1, row=row_no).value
                    unit_name = re.sub('\t+', '', unit_name_unformatted)
                    invigilator = ws.cell(column=int(column_no)+2, row=row_no).value
                    venue = ws.cell(column=int(column_no)+3, row=row_no).value
                    time = ws.cell(column=column_no, row=2).value

                    # Get the date for specific exams
                    for row in ws.iter_rows(min_row=date_start_row, max_row=row_no, min_col=1, max_col=1): 
                        for cell in row:
                            if cell.value is not None and "DAY" in cell.value:
                                date = str(ws.cell(column=1, row=int(cell.row)+1).value)+' '+str(cell.value)

                    # TODO: Send extracted data to db

                    db_year = Year.objects.get(year=year)
                    db_sem = Semester.objects.get(Q(year=db_year.year),Q(sem=sem))

                    exam = Exam(
                        semester = db_sem,
                        timetable_id = timetable_id,
                        exam_no = count,
                        unit_name = unit_name,
                        invigilator = invigilator,
                        venue = venue,
                        date = date,
                        time = time
                    )

                    exam.save()
                    
                    count += 1
